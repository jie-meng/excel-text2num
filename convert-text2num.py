import os
import os.path

from openpyxl import load_workbook

PURE_NUM_CHARACTERS = '1234567890'
NUM_CHARACTERS = '1234567890.,'

def scriptPath():
    return os.path.dirname(os.path.abspath(__file__))

def is_int(s):
    try:
        int(s)
        return True
    except ValueError:
        return None

def is_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return None

def mayBeNumber(str):
    point_count = 0
    for c in str:
        if NUM_CHARACTERS.find(c) >= 0:
            if c == '.':
                point_count += 1
        else:
            return False

    return point_count < 2

def removeCommas(str):
    return str.replace(',', '')

def isID(str):
    if len(str) >= 12:
        idx = 0
        for c in str:
            if PURE_NUM_CHARACTERS.find(c) < 0:
                return False

            idx += 1
            if (idx >= 12):
                break

        if str.startswith('18') or str.startswith('62') or str.startswith('12') or str.startswith('13') or str.startswith('05') or str.startswith('08'):
            return True

    return False

def process_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if type(cell.value) == str:
                if mayBeNumber(cell.value) and (not isID(cell.value)):
                    processed_str = removeCommas(cell.value)
                    if is_int(processed_str) != None:
                        print('cell {0} converted to int {1}'.format(cell, processed_str))
                        cell.value = int(processed_str)
                    elif is_float(processed_str) != None:
                        print('cell {0} converted float {1}'.format(cell, processed_str))
                        cell.value = float(processed_str)

def process_workbook(source, dest):
    wb = load_workbook(source)

    for sheet in wb.worksheets:
        process_sheet(sheet)

    wb.save(dest)

if __name__ == "__main__":
    print('please input source xlsx file path:')
    source = input().strip()

    output_dir = '{0}/output'.format(scriptPath())
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    outputfile = '{0}/output/{1}'.format(scriptPath(), os.path.basename(source))

    process_workbook(source, outputfile)

    print('\noutput to {0}'.format(outputfile))
    print('\nPress any key to continue ...')
    input()

